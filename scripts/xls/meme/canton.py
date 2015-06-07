#!/usr/local/bin/python3


def printFun(list):
  for name in list[:-1]:
    print('matu_ecole == "' + name + '" | ///')
  print('matu_ecole == "' + list[-1] + '"')


from openpyxl import load_workbook
wb2 = load_workbook('listeEcoles.xlsx', data_only=True)
ws = wb2['Sheet1']

# print(ws.cell(row = 2, column = 1).value)
ret = {}
for x in range(1, 107):
  index =  ws.cell(row = x, column = 7).value
  if index not in ret:
    ret[index] = []
  ret[index].append(ws.cell(row = x, column = 1).value)

print("gen canton = .")

strMain = "replace canton = {0} if ///"
strIf = """
"""
for key in ret:
  print(strMain.format(key))
  printFun(ret[key])