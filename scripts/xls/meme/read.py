#!/usr/local/bin/python3


def printFun(list):
  for name in list[:-1]:
    print('cours == "' + name + '" | ///')
  print('cours == "' + list[-1] + '"')


from openpyxl import load_workbook
wb2 = load_workbook('listeCours.xlsx')
ws = wb2['Sheet1']

# print ws.cell(row = 2, column = 1).value

cat0 = []
cat1 = []
cat2 = []


for x in range(2, 187):
  if ws.cell(row = x, column = 2).value == 0:
    cat0.append(ws.cell(row = x, column = 3).value)
  elif ws.cell(row = x, column = 2).value == 1:
    cat1.append(ws.cell(row = x, column = 3).value)
  elif ws.cell(row = x, column = 2).value == 2:
    cat2.append(ws.cell(row = x, column = 3).value)


print('gen sci = .\nreplace sci = 0 if ///')
printFun(cat0)

print('replace sci = 1 if ///')
printFun(cat1)

print('replace sci = 2 if ///')
printFun(cat2)



# mixte = []
# notMixte = []
# for x in range(2, 186):
#   if ws.cell(row = x, column = 2).value == 2:
#     mixte.append(ws.cell(row = x, column = 3).value)
#   elif ws.cell(row = x, column = 2).value == 0 or ws.cell(row = x, column = 2).value == 1:
#     notMixte.append(ws.cell(row = x, column = 3).value)


# print('gen mixte = .\nreplace mixte = 1 if ///')
# printFun(mixte)

# print('replace mixte = 0 if ///')
# printFun(notMixte)


print ('gen year = .\n')
year = [[], [], []]
for x in range(2, 186):
  if not(ws.cell(row = x, column = 1).value is None):
    year[int(ws.cell(row = x, column = 1).value)-1].append(ws.cell(row = x, column = 3).value)

print ('replace year = 1 if ///')
printFun(year[0])

print ('replace year = 2 if ///')
printFun(year[1])

print ('replace year = 3 if ///')
printFun(year[2])