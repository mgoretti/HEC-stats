#!/usr/local/bin/python3


def printFun(list):
  for name in list[:-1]:
    print('cours == "' + name + '" | ///')
  print('cours == "' + list[-1] + '"')


from openpyxl import load_workbook
wb2 = load_workbook('labelCours.xlsx')
ws = wb2['Sheet1']

# print(ws.cell(row = 2, column = 1).value)

final = []
for col in range(7, 11):
  ret = []
  for x in range(1, 186):
    if ws.cell(row = x, column = col).value == 1:
      ret.append(x)
  final.append(ret)
  print('la')

print(final)