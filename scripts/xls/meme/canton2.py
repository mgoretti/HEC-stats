#!/usr/local/bin/python3


def printFun(list):
  for val in list[:-1]:
    if val == 0:
      print('NA' + ', ', end="")
    else:
      print(str(val) + ', ', end="")
  print(str(list[-1]) + '\n)', end="")


from openpyxl import load_workbook
wb2 = load_workbook('listeEcoles.xlsx', data_only=True)
ws = wb2['Sheet1']

# print(ws.cell(row = 2, column = 1).value)
ret = [0] * 26
for x in range(2, 28):
  index = ws.cell(row = x, column = 18).value
  if index != None:
    ret[index-1]=(ws.cell(row = x, column = 19).value)

print("data <- c(")
printFun(ret)
